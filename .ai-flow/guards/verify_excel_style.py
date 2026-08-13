#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""更新後の Excel を style YAML（確定済みの書式の正）と突合し、書式違反を検出する。

使い方:
    python verify_excel_style.py --style <確定済み.style.yaml> <対象.xlsx> [<対象2.xlsx> ...]

検査項目（style YAML の定義に応じて実施）:
  validation.allowed_fonts : 許可フォント以外の使用を ERROR（例: ＭＳ ゴシック文書への游ゴシック混入）
  validation.allowed_sizes : （任意）許可サイズ以外の使用を ERROR
  patterns.*               : anchor ラベルで表を特定し、ヘッダ行の列構成が columns と一致するか検査
                             （anchor が見つからないシートはスキップ。座標でなくラベルで位置決めする）

[7]ドキュメント反映の完了前に実行し、ERROR は解消してからゲートに進む
（誤検知と考える場合も独断で style YAML を書き換えず、ユーザーに確認する）。

exit code: 0=違反なし / 1=ERROR あり / 3=実行エラー（依存不足・ファイル不正等）
"""
import argparse
import sys

MAX_REPORT_PER_CHECK = 30


def fail(msg):
    sys.stderr.write(msg + '\n')
    sys.exit(3)


def check_fonts(wb, allowed_fonts, allowed_sizes, findings):
    over = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                name = cell.font.name or '(default)'
                size = float(cell.font.size) if cell.font.size else None
                bad = None
                if allowed_fonts and name not in allowed_fonts:
                    bad = '許可外フォント "%s"' % name
                elif allowed_sizes and size is not None and size not in allowed_sizes:
                    bad = '許可外サイズ %g' % size
                if bad:
                    if len(findings) < MAX_REPORT_PER_CHECK:
                        findings.append('ERROR %s!%s: %s' % (ws.title, cell.coordinate, bad))
                    else:
                        over += 1
    if over:
        findings.append('ERROR ...他 %d 件（フォント/サイズ違反）' % over)


def check_patterns(wb, patterns, findings):
    for pat_name, pat in (patterns or {}).items():
        anchor = str(pat.get('anchor') or '').strip()
        columns = [str(c) for c in (pat.get('columns') or [])]
        if not anchor or not columns:
            continue
        offset = pat.get('header_offset') or {}
        d_row, d_col = int(offset.get('row', 1)), int(offset.get('col', 0))
        found = False
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None or str(cell.value).strip() != anchor:
                        continue
                    found = True
                    hr, hc = cell.row + d_row, cell.column + d_col
                    actual = [ws.cell(row=hr, column=hc + i).value for i in range(len(columns))]
                    actual = ['' if v is None else str(v).strip() for v in actual]
                    if actual != columns:
                        findings.append(
                            'ERROR %s!%s 付近: 表「%s」のヘッダが規約と不一致 期待=%s 実際=%s'
                            % (ws.title, cell.coordinate, pat_name, columns, actual))
        if not found:
            # anchor 自体が無いのは違反にしない（その表を含まないドキュメントもある）
            findings.append('INFO 表「%s」（anchor "%s"）は対象内に見つからないためスキップ'
                            % (pat_name, anchor))


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--style', required=True, help='確定済み style YAML のパス')
    ap.add_argument('targets', nargs='+', help='検査対象の Excel ファイル（.xlsx/.xlsm）')
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        fail('openpyxl がありません: pip install openpyxl pyyaml')
    try:
        import yaml
    except ImportError:
        fail('PyYAML がありません: pip install openpyxl pyyaml')

    try:
        with open(args.style, encoding='utf-8') as f:
            spec = yaml.safe_load(f) or {}
    except Exception as e:
        fail('style YAML を読めません: %s (%s)' % (args.style, e))

    validation = spec.get('validation') or {}
    allowed_fonts = set(validation.get('allowed_fonts') or [])
    allowed_sizes = {float(s) for s in (validation.get('allowed_sizes') or [])}
    patterns = spec.get('patterns') or {}

    errors = 0
    for target in args.targets:
        try:
            wb = openpyxl.load_workbook(target, data_only=False)
        except Exception as e:
            fail('Excel を開けません: %s (%s)' % (target, e))
        findings = []
        check_fonts(wb, allowed_fonts, allowed_sizes, findings)
        check_patterns(wb, patterns, findings)
        n_err = sum(1 for x in findings if x.startswith('ERROR'))
        errors += n_err
        print('== %s: ERROR %d 件 ==' % (target, n_err))
        for x in findings:
            print('  ' + x)

    return 1 if errors else 0


if __name__ == '__main__':
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding='utf-8')
        except Exception:
            pass
    sys.exit(main())

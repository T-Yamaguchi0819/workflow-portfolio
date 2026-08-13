#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""既存 Excel ドキュメントから書式カタログ（style YAML の下書き）を抽出する。

使い方:
    python extract_excel_style.py <サンプル.xlsx> [-o 出力.style.yaml] [--top N]

- 依存: openpyxl・PyYAML（pip install openpyxl pyyaml）
- セル書式の組み合わせ（フォント・サイズ・太字・塗り・罫線・配置・表示形式）を
  出現頻度順に集計し、上位 N 件を style 候補として出力する
- 出力は**下書き**。人間がレビューして styles の命名・validation（許可フォント等）・
  patterns（表のアンカーと列構成）を確定してから docs-style/ に置く（確定後が「書式の正」）
- 図形・グラフ・条件付き書式は抽出対象外（生成側は雛形 .xlsx で保全する方針）

exit code: 0=成功 / 3=実行エラー（依存不足・ファイル不正等）
"""
import argparse
import sys
from collections import Counter, defaultdict


def fail(msg):
    sys.stderr.write(msg + '\n')
    sys.exit(3)


def fill_sig(fill):
    try:
        if fill is not None and fill.patternType == 'solid':
            rgb = fill.fgColor.rgb
            if isinstance(rgb, str) and len(rgb) >= 6:
                return rgb[-6:]
    except Exception:
        pass
    return 'none'


def border_sig(border):
    edges = [e for e in ('top', 'bottom', 'left', 'right')
             if getattr(border, e, None) is not None and getattr(border, e).style]
    return '+'.join(edges) if edges else 'none'


def cell_sig(cell):
    f = cell.font
    return (
        f.name or '(default)',
        float(f.size) if f.size else None,
        bool(f.bold),
        fill_sig(cell.fill),
        border_sig(cell.border),
        cell.alignment.horizontal or 'default',
        cell.number_format or 'General',
    )


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('sample', help='書式抽出元の Excel ファイル（.xlsx/.xlsm）')
    ap.add_argument('-o', '--output', default=None,
                    help='出力先 YAML（省略時: <サンプル名>.style.yaml）')
    ap.add_argument('--top', type=int, default=20, help='style 候補の上限数（既定 20）')
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        fail('openpyxl がありません: pip install openpyxl pyyaml')
    try:
        import yaml
    except ImportError:
        fail('PyYAML がありません: pip install openpyxl pyyaml')

    try:
        wb = openpyxl.load_workbook(args.sample, data_only=False)
    except Exception as e:
        fail('Excel を開けません: %s (%s)' % (args.sample, e))

    counts = Counter()
    examples = defaultdict(list)
    sheets = []
    for ws in wb.worksheets:
        sheets.append({'name': ws.title, 'max_row': ws.max_row, 'max_col': ws.max_column})
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                sig = cell_sig(cell)
                counts[sig] += 1
                if len(examples[sig]) < 3:
                    examples[sig].append('%s!%s' % (ws.title, cell.coordinate))

    styles = {}
    for i, (sig, n) in enumerate(counts.most_common(args.top), start=1):
        name, size, bold, fill, border, align, numfmt = sig
        styles['style_%02d' % i] = {
            'font': name, 'size': size, 'bold': bold, 'fill': fill,
            'border': border, 'align': align, 'number_format': numfmt,
            'count': n, 'examples': examples[sig],
        }

    fonts = sorted({sig[0] for sig in counts})
    data = {
        'source': args.sample,
        'sheets': sheets,
        'styles': styles,
        # patterns: 表をラベルで特定して列構成を検査する定義（人間が追記して確定する）
        'patterns': {},
        'validation': {
            'allowed_fonts': fonts,   # 混入を許さないフォントを人間が削って確定する
        },
    }

    header = (
        '# 書式カタログ（extract_excel_style.py による下書き）\n'
        '# 人間がレビューして確定してから docs-style/ に置くこと:\n'
        '#  - styles: 意味のある名前に変更し、不要な候補を削る（例: style_01 → 表ヘッダ）\n'
        '#  - validation.allowed_fonts: 許可しないフォントを削る（混入検出に使われる）\n'
        '#  - patterns: 検査したい表を追記する。例:\n'
        '#      patterns:\n'
        '#        改訂履歴:\n'
        '#          anchor: "改訂履歴"          # セル座標でなくラベルで位置決め\n'
        '#          header_offset: {row: 1, col: 0}\n'
        '#          columns: [版数, 日付, 変更者, 変更内容]\n'
        '# 検査は guards/verify_excel_style.py が行う\n'
    )
    out = args.output or (args.sample.rsplit('.', 1)[0] + '.style.yaml')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(header)
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False, default_flow_style=False)
    print('書式カタログの下書きを出力しました: %s（styles %d 件・フォント %d 種）'
          % (out, len(styles), len(fonts)))
    return 0


if __name__ == '__main__':
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding='utf-8')
        except Exception:
            pass
    sys.exit(main())
